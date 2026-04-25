import openpyxl
from openpyxl.utils import get_column_letter
from src.safwanbuddy.core import logger
import os

class ExcelGenerator:
    def __init__(self):
        self.output_dir = "output/spreadsheets"
        os.makedirs(self.output_dir, exist_ok=True)

    def create_spreadsheet(self, data: list, filename: str = "spreadsheet.xlsx", sheet_name: str = "Sheet1"):
        """
        Creates an Excel spreadsheet with professional formatting.
        data is a list of lists (rows).
        """
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            # Styling
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            for r_idx, row in enumerate(data, 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = thin_border
                    if r_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_alignment

            # Auto-adjust column widths
            for column_cells in ws.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = max_length + 5

            path = os.path.join(self.output_dir, filename)
            wb.save(path)
            logger.info(f"Excel spreadsheet saved to {path}")
            return path
        except Exception as e:
            logger.error(f"Error creating Excel spreadsheet: {e}")
            return None

excel_generator = ExcelGenerator()
