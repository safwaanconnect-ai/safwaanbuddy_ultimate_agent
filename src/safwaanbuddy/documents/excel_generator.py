"""Excel spreadsheet generation using openpyxl."""

import logging
from pathlib import Path
from typing import List, Dict, Any, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not available")

from ..core.events import EventBus, EventType
from ..core.config import ConfigManager


class ExcelGenerator:
    """Generate Excel spreadsheets."""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.config = ConfigManager()
        self.event_bus = EventBus()
        
        self.workbook: Optional[Workbook] = None
        self.worksheet = None
    
    def create_workbook(self, sheet_name: str = "Sheet1") -> bool:
        """Create new workbook.
        
        Args:
            sheet_name: Name of first sheet
            
        Returns:
            True if successful
        """
        if not OPENPYXL_AVAILABLE:
            self.logger.error("openpyxl not available")
            return False
        
        try:
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = sheet_name
            self.logger.info("Created new Excel workbook")
            return True
        except Exception as e:
            self.logger.error(f"Failed to create workbook: {e}")
            return False
    
    def write_data(self, data: List[List[Any]], start_row: int = 1, start_col: int = 1) -> bool:
        """Write data to worksheet.
        
        Args:
            data: Data to write (list of rows)
            start_row: Starting row (1-based)
            start_col: Starting column (1-based)
            
        Returns:
            True if successful
        """
        if not self.worksheet:
            return False
        
        try:
            for i, row in enumerate(data):
                for j, value in enumerate(row):
                    self.worksheet.cell(row=start_row + i, column=start_col + j, value=value)
            return True
        except Exception as e:
            self.logger.error(f"Failed to write data: {e}")
            return False
    
    def save_workbook(self, filepath: Path) -> bool:
        """Save workbook to file.
        
        Args:
            filepath: Output file path
            
        Returns:
            True if successful
        """
        if not self.workbook:
            return False
        
        try:
            filepath = Path(filepath)
            filepath.parent.mkdir(parents=True, exist_ok=True)
            
            self.workbook.save(filepath)
            
            self.event_bus.emit(EventType.DOCUMENT_SAVED, {
                "type": "excel",
                "filepath": str(filepath)
            })
            
            self.logger.info(f"Workbook saved: {filepath}")
            return True
        except Exception as e:
            self.logger.error(f"Failed to save workbook: {e}")
            return False
